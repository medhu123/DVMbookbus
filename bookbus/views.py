import logging
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib.auth.models import User
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView, TemplateView, View
from .models import Bus, Booking, BusStop, Seat, Stop
from .forms import BookingSeatForm, FilterForm, BusForm, StopForm, BusStopForm, SeatSelectionForm, PassengerInfoForm
from django.db.models import Q, F, Subquery, OuterRef, Prefetch, Exists, BooleanField, Value
import datetime
from django.utils import timezone
from django.forms import ValidationError
from django.contrib import messages
from django.db import transaction
from django.http import HttpResponse, HttpResponseForbidden
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

logger = logging.getLogger(__name__)

def home(request):
    form = FilterForm(request.GET or None)
    buses = Bus.objects.prefetch_related(
        Prefetch('bus_stops', queryset=BusStop.objects.select_related('stop').order_by('stop_order'))
    ).all()
    
    if form.is_valid():
        start_stop = form.cleaned_data.get('journey_start')
        end_stop = form.cleaned_data.get('journey_end')
        travel_date = form.cleaned_data.get('travel_date')
        
        # Stop order filtering
        if start_stop and end_stop:
            start_pos = Subquery(
                BusStop.objects.filter(
                    bus=OuterRef('pk'),
                    stop=start_stop
                ).values('stop_order')[:1]
            )
            end_pos = Subquery(
                BusStop.objects.filter(
                    bus=OuterRef('pk'),
                    stop=end_stop
                ).values('stop_order')[:1]
            )
            buses = buses.annotate(
                start_pos=start_pos,
                end_pos=end_pos
            ).filter(
                start_pos__isnull=False,
                end_pos__isnull=False,
                start_pos__lt=F('end_pos')
            ).distinct()
        elif start_stop:
            buses = buses.filter(bus_stops__stop=start_stop)
        elif end_stop:
            buses = buses.filter(bus_stops__stop=end_stop)
        
        # Date filtering - now shows both recurring and date-range matches
        if travel_date:
            if isinstance(travel_date, str):
                travel_date = datetime.datetime.strptime(travel_date, '%Y-%m-%d').date()
            
            buses = [bus for bus in buses if bus.runs_on_date(travel_date)]
    
    context = {
        'form': form,
        'buses': buses,
        'today': datetime.datetime.now().date(),
        'start_cities': Stop.objects.all(),
        'end_cities': Stop.objects.all(),
        'is_paginated': True,
        'page_obj': buses
    }
    return render(request, 'bookbus/home.html', context)


class BusListView(ListView):
    model = Bus
    template_name = 'bookbus/home.html'
    context_object_name = 'buses'
    ordering = ['-start_time']
    paginate_by = 5


class UserBusListView(ListView):
    model = Bus
    template_name = 'bookbus/user_buses.html'
    ordering = ['-start_time']
    context_object_name = 'buses'
    paginate_by = 5

    def get_queryset(self):
        user = get_object_or_404(User, username=self.kwargs.get('username'))
        return Bus.objects.filter(travels=user).order_by('-start_time')


class UserBookingListView(ListView):
    model = Booking
    template_name = 'bookbus/booked_buses.html'
    context_object_name = 'bookings'
    paginate_by = 5

    def get_queryset(self):
        user = get_object_or_404(User, username=self.kwargs.get('username'))
        return user.booking_set.all()


class BusDetailView(DetailView):
    model = Bus

class BusCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = Bus
    form_class = BusForm
    template_name = 'bookbus/bus_form.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['stop_form'] = BusStopForm()
        context['all_stops'] = Stop.objects.all()
        return context

    def form_valid(self, form):
        self.object = form.save(commit=False)
        self.object.travels = self.request.user
        
        # Get stops and times from form data
        stop_order = self.request.POST.get('stop_order', '').split(',')
        stop_times = self.request.POST.get('stop_times', '').split(',')
        next_day_flags = self.request.POST.get('next_day_flags', '').split(',')
        
        stop_order = [stop_id for stop_id in stop_order if stop_id]
        stop_times = [time for time in stop_times if time]
        next_day_flags = [flag for flag in next_day_flags if flag]
        
        if len(stop_order) < 2:
            form.add_error(None, "Please select at least 2 stops")
            return self.form_invalid(form)
            
        if len(stop_order) != len(stop_times) or len(stop_order) != len(next_day_flags):
            form.add_error(None, "Invalid stop data")
            return self.form_invalid(form)
            
        try:
            # Set journey start/end
            self.object.journey_start = Stop.objects.get(id=stop_order[0])
            self.object.journey_end = Stop.objects.get(id=stop_order[-1])
            self.object.save()
            
            # Create BusStop relationships with times
            for order, (stop_id, time_str, next_day) in enumerate(zip(stop_order, stop_times, next_day_flags), start=1):
                try:
                    time_obj = datetime.datetime.strptime(time_str, "%H:%M").time()
                except ValueError:
                    form.add_error(None, f"Invalid time format for stop {stop_id}")
                    return self.form_invalid(form)
                    
                BusStop.objects.create(
                    bus=self.object,
                    stop_id=stop_id,
                    stop_order=order,
                    arrival_time=time_obj,
                    is_next_day=next_day == '1'
                )
            
            # Create Seat objects
            self.create_seats(
                general_count=form.cleaned_data['general_count'],
                sleeper_count=form.cleaned_data['sleeper_count'],
                luxury_count=form.cleaned_data['luxury_count'],
                general_fare=form.cleaned_data['general_fare'],
                sleeper_fare=form.cleaned_data['sleeper_fare'],
                luxury_fare=form.cleaned_data['luxury_fare']
            )
                
            return redirect('bus-detail', pk=self.object.pk)
            
        except Stop.DoesNotExist:
            form.add_error(None, "One or more selected stops no longer exist")
            return self.form_invalid(form)
    
    def create_seats(self, general_count, sleeper_count, luxury_count, general_fare, sleeper_fare, luxury_fare):
        # Create General seats
        for i in range(1, general_count + 1):
            Seat.objects.create(
                bus=self.object,
                name=f"G{i}",
                seat_class="General",
                fare=general_fare
            )
        
        # Create Sleeper seats
        for i in range(1, sleeper_count + 1):
            Seat.objects.create(
                bus=self.object,
                name=f"S{i}",
                seat_class="Sleeper",
                fare=sleeper_fare
            )
        
        # Create Luxury seats
        for i in range(1, luxury_count + 1):
            Seat.objects.create(
                bus=self.object,
                name=f"L{i}",
                seat_class="Luxury",
                fare=luxury_fare
            )
    
    def test_func(self):
        return self.request.user.groups.filter(name='BusAdmin').exists()


class BusUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = Bus
    form_class = BusForm
    template_name = 'bookbus/bus_form.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        current_stops = self.object.bus_stops.order_by('stop_order')
        
        # Prepare initial data for stops, times and next_day flags
        stop_ids = []
        stop_times = []
        next_day_flags = []
        
        for bs in current_stops:
            stop_ids.append(str(bs.stop.id))
            stop_times.append(bs.arrival_time.strftime("%H:%M"))
            next_day_flags.append('1' if bs.is_next_day else '0')
            
        initial_data = {
            'stop_order': ','.join(stop_ids),
            'stop_times': ','.join(stop_times),
            'next_day_flags': ','.join(next_day_flags)
        }
        
        context['stop_form'] = BusStopForm(initial=initial_data)
        context['all_stops'] = Stop.objects.all()
        context['current_stops'] = current_stops
        return context

    def form_valid(self, form):
        self.object = form.save()
        
        # Clear existing stops
        self.object.bus_stops.all().delete()
        
        # Get new stop data
        stop_order = self.request.POST.get('stop_order', '').split(',')
        stop_times = self.request.POST.get('stop_times', '').split(',')
        next_day_flags = self.request.POST.get('next_day_flags', '').split(',')
        
        stop_order = [stop_id for stop_id in stop_order if stop_id]
        stop_times = [time for time in stop_times if time]
        next_day_flags = [flag for flag in next_day_flags if flag]
        
        if len(stop_order) < 2:
            form.add_error(None, "Please select at least 2 stops")
            return self.form_invalid(form)
            
        if len(stop_order) != len(stop_times) or len(stop_order) != len(next_day_flags):
            form.add_error(None, "Invalid stop data")
            return self.form_invalid(form)
            
        try:
            # Set journey start/end
            self.object.journey_start = Stop.objects.get(id=stop_order[0])
            self.object.journey_end = Stop.objects.get(id=stop_order[-1])
            self.object.save()
            
            # Create BusStop relationships with times
            for order, (stop_id, time_str, next_day) in enumerate(zip(stop_order, stop_times, next_day_flags), start=1):
                try:
                    time_obj = datetime.datetime.strptime(time_str, "%H:%M").time()
                except ValueError:
                    form.add_error(None, f"Invalid time format for stop {stop_id}")
                    return self.form_invalid(form)
                    
                BusStop.objects.create(
                    bus=self.object,
                    stop_id=stop_id,
                    stop_order=order,
                    arrival_time=time_obj,
                    is_next_day=next_day == '1'
                )
            
            # Recreate seats if configuration changed
            if self.seat_config_changed(form.cleaned_data):
                self.object.seats.all().delete()
                self.create_seats(
                    general_count=form.cleaned_data['general_count'],
                    sleeper_count=form.cleaned_data['sleeper_count'],
                    luxury_count=form.cleaned_data['luxury_count'],
                    general_fare=form.cleaned_data['general_fare'],
                    sleeper_fare=form.cleaned_data['sleeper_fare'],
                    luxury_fare=form.cleaned_data['luxury_fare']
                )
                
            return redirect('bus-detail', pk=self.object.pk)
            
        except Stop.DoesNotExist:
            form.add_error(None, "One or more selected stops no longer exist")
            return self.form_invalid(form)
    
    def seat_config_changed(self, cleaned_data):
        current_counts = {
            'general': self.object.seats.filter(seat_class="General").count(),
            'sleeper': self.object.seats.filter(seat_class="Sleeper").count(),
            'luxury': self.object.seats.filter(seat_class="Luxury").count()
        }
        current_fares = {
            'general': self.object.seats.filter(seat_class="General").first().fare if self.object.seats.filter(seat_class="General").exists() else 0,
            'sleeper': self.object.seats.filter(seat_class="Sleeper").first().fare if self.object.seats.filter(seat_class="Sleeper").exists() else 0,
            'luxury': self.object.seats.filter(seat_class="Luxury").first().fare if self.object.seats.filter(seat_class="Luxury").exists() else 0
        }
        
        return (
            cleaned_data['general_count'] != current_counts['general'] or
            cleaned_data['sleeper_count'] != current_counts['sleeper'] or
            cleaned_data['luxury_count'] != current_counts['luxury'] or
            cleaned_data['general_fare'] != current_fares['general'] or
            cleaned_data['sleeper_fare'] != current_fares['sleeper'] or
            cleaned_data['luxury_fare'] != current_fares['luxury']
        )

    def create_seats(self, general_count, sleeper_count, luxury_count, general_fare, sleeper_fare, luxury_fare):
        # Create General seats
        for i in range(1, general_count + 1):
            Seat.objects.create(
                bus=self.object,
                name=f"G{i}",
                seat_class="General",
                fare=general_fare
            )
        
        # Create Sleeper seats
        for i in range(1, sleeper_count + 1):
            Seat.objects.create(
                bus=self.object,
                name=f"S{i}",
                seat_class="Sleeper",
                fare=sleeper_fare
            )
        
        # Create Luxury seats
        for i in range(1, luxury_count + 1):
            Seat.objects.create(
                bus=self.object,
                name=f"L{i}",
                seat_class="Luxury",
                fare=luxury_fare
            )

    def test_func(self):
        bus = self.get_object()
        return self.request.user == bus.travels

class BusDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    model = Bus
    success_url = '/'

    def test_func(self):
        bus = self.get_object()
        return self.request.user == bus.travels
        
def bus_book(request, pk):
    bus = get_object_or_404(Bus, pk=pk)
    
    # Initialize booking data from session
    booking_data = request.session.get('booking_data', {
        'selected_seats': [],
        'travel_date': None,
    })

    if request.method == 'POST':
        if 'update_date' in request.POST:
            # Handle date change
            travel_date_str = request.POST.get('travel_date')
            if not travel_date_str:
                return redirect('bus-book', pk=bus.pk)
                
            try:
                travel_date = datetime.datetime.strptime(travel_date_str, '%Y-%m-%d').date()
                
                # Validate date
                if travel_date < datetime.date.today():
                    messages.error(request, "Travel date cannot be in the past")
                elif not bus.runs_on_date(travel_date):
                    messages.error(request, "Bus doesn't operate on selected date")
                else:
                    # Update session with new date and clear selections
                    booking_data['travel_date'] = travel_date_str
                    booking_data['selected_seats'] = []
                    request.session['booking_data'] = booking_data
                    request.session.modified = True
                    
            except ValueError:
                messages.error(request, "Invalid date format")
            return redirect('bus-book', pk=bus.pk)
        
        if 'select_seats' in request.POST:
            # Validate travel date first
            if not booking_data.get('travel_date'):
                messages.error(request, "Please select a valid travel date first")
                return redirect('bus-book', pk=bus.pk)
                
            try:
                travel_date = datetime.datetime.strptime(booking_data['travel_date'], '%Y-%m-%d').date()
            except ValueError:
                messages.error(request, "Invalid travel date in session")
                return redirect('bus-book', pk=bus.pk)
            
            # Get selected seats from POST data - ensure we only get unique seats
            selected_seats = list(set(request.POST.getlist('seats')))
            
            if not selected_seats:
                messages.error(request, "Please select at least one seat")
                return redirect('bus-book', pk=bus.pk)
            
            # Convert to integers and validate they exist
            valid_seat_ids = []
            for seat_id in selected_seats:
                try:
                    seat_id_int = int(seat_id)
                    if Seat.objects.filter(id=seat_id_int, bus=bus).exists():
                        valid_seat_ids.append(seat_id_int)
                except (ValueError, TypeError):
                    continue
            
            if not valid_seat_ids:
                messages.error(request, "Invalid seat selection")
                return redirect('bus-book', pk=bus.pk)
            
            # Check seat availability
            unavailable_seats = Booking.objects.filter(
                seat_id__in=valid_seat_ids,
                travel_date=travel_date,
                status='Confirmed'  # Only check confirmed bookings
            ).values_list('seat_id', flat=True)
            
            if unavailable_seats.exists():
                messages.error(request, f"Seat(s) {', '.join(map(str, unavailable_seats))} are no longer available")
                return redirect('bus-book', pk=bus.pk)
            
            # Update session with selected seats
            booking_data['selected_seats'] = valid_seat_ids
            request.session['booking_data'] = booking_data
            request.session.modified = True
            
            messages.success(request, f"Selected {len(valid_seat_ids)} seat(s). Please enter passenger details.")
            return redirect('bus-book', pk=bus.pk)
        
        elif 'confirm_booking' in request.POST:
            if not booking_data.get('travel_date') or not booking_data.get('selected_seats'):
                messages.error(request, "Invalid booking data")
                return redirect('bus-book', pk=bus.pk)
                
            try:
                travel_date = datetime.datetime.strptime(booking_data['travel_date'], '%Y-%m-%d').date()
                selected_seats = booking_data['selected_seats']
            except:
                messages.error(request, "Invalid booking data")
                return redirect('bus-book', pk=bus.pk)

            # Calculate total cost
            seats = Seat.objects.filter(id__in=selected_seats, bus=bus)
            total_cost = sum(seat.fare for seat in seats)

            profile = request.user.profile
            
            # Check if user has enough coins
            if profile.coins < total_cost:
                messages.error(request, f"You don't have enough coins. Needed: {total_cost}, Available: {profile.coins}")
                return redirect('bus-book', pk=bus.pk)
            
            # Only validate for the actual selected seats
            passenger_errors = []
            for i in range(len(selected_seats)):
                name = request.POST.get(f'passenger_name_{i}', '').strip()
                email = request.POST.get(f'passenger_email_{i}', '').strip()
                
                if not name:
                    passenger_errors.append(f"Passenger {i+1}: Name is required")
                if not email or '@' not in email:
                    passenger_errors.append(f"Passenger {i+1}: Valid email is required")
            
            if passenger_errors:
                for error in passenger_errors:
                    messages.error(request, error)
                return redirect('bus-book', pk=bus.pk)
            
            # Process booking in transaction
            with transaction.atomic():
                successful_bookings = 0
                unavailable_seats = []
                for i, seat_id in enumerate(selected_seats):
                    try:
                        seat = Seat.objects.select_for_update().get(pk=seat_id, bus=bus)
                        
                        # Double-check availability
                        if Booking.objects.filter(seat=seat, travel_date=travel_date).exists():
                            messages.warning(request, f"Seat {seat.name} was already booked")
                            continue
                        
                        # Get boarding and destination points
                        try:
                            start_stop = bus.bus_stops.get(id=request.POST.get(f'start_stop_{i}'))
                            end_stop = bus.bus_stops.get(id=request.POST.get(f'end_stop_{i}'))
                        except (ValueError, BusStop.DoesNotExist):
                            messages.error(request, f"Invalid stops for seat {seat.name}")
                            continue
                        
                        # Create booking
                        Booking.objects.create(
                            bus=bus,
                            customer=request.user,
                            seat=seat,
                            start_stop=start_stop.stop,
                            end_stop=end_stop.stop,
                            travel_date=travel_date,
                            passenger_name=request.POST.get(f'passenger_name_{i}'),
                            passenger_email=request.POST.get(f'passenger_email_{i}'),
                            status='Confirmed'
                        )
                        successful_bookings += 1
                        
                    except Exception as e:
                        messages.error(request, f"Error booking seat {seat_id}: {str(e)}")
                        continue
                
                if successful_bookings > 0:
                     # Deduct coins only for successful bookings
                    actual_cost = sum(seat.fare for seat in seats.filter(id__in=[s for s in selected_seats if s not in unavailable_seats]))
                    profile.coins -= actual_cost
                    profile.save()

                    del request.session['booking_data']
                    messages.success(request, f"Successfully booked {successful_bookings} seat(s)! {actual_cost} coins deducted.")
                    return redirect('booked-buses', username=request.user.username)
                else:
                    messages.error(request, "No seats were booked due to errors")
                    return redirect('bus-book', pk=bus.pk)

    # Prepare context for GET requests or after POST processing
    travel_date = None
    if booking_data.get('travel_date'):
        try:
            travel_date = datetime.datetime.strptime(booking_data['travel_date'], '%Y-%m-%d').date()
        except (ValueError, TypeError):
            booking_data['travel_date'] = None
            request.session['booking_data'] = booking_data
            request.session.modified = True

    # Get seats with availability status
    seats = bus.seats.annotate(
        is_booked=Exists(
            Booking.objects.filter(
                seat=OuterRef('pk'),
                travel_date=travel_date,
                status='Confirmed'
            )
        ) if travel_date else Value(False, output_field=BooleanField())
    ).order_by('seat_class', 'name')

    # Natural sorting function
    def natural_sort_key(seat):
        name = seat.name
        alpha = ''
        num = ''
        for char in name:
            if char.isdigit():
                num += char
            else:
                alpha += char
        return (seat.seat_class, alpha.lower(), int(num) if num else 0)

    seats = sorted(seats, key=natural_sort_key)

    # Get selected seat objects for passenger info
    selected_seat_objects = Seat.objects.filter(
        id__in=booking_data.get('selected_seats', []),
        bus=bus
    )

    context = {
        'bus': bus,
        'travel_date': travel_date,
        'min_date': datetime.date.today(),
        'max_date': bus.end_time.date(),
        'seats_by_class': {
            'General': [s for s in seats if s.seat_class == 'General'],
            'Sleeper': [s for s in seats if s.seat_class == 'Sleeper'],
            'Luxury': [s for s in seats if s.seat_class == 'Luxury'],
        },
        'selected_seats': booking_data.get('selected_seats', []),
        'selected_seat_objects': selected_seat_objects,
        'bus_stops': bus.bus_stops.select_related('stop').order_by('stop_order')
    }
    
    return render(request, 'bookbus/bus_book.html', context)
    
    
def passenger_info(request, pk):
    bus = get_object_or_404(Bus, pk=pk)
    
    if 'selected_seats' not in request.session:
        return redirect('bus-book', pk=bus.pk)
    
    try:
        # Convert travel_date to date object if it's a string
        travel_date = request.session['travel_date']
        if isinstance(travel_date, str):
            travel_date = datetime.datetime.strptime(travel_date, '%Y-%m-%d').date()
        
        seat_count = len(request.session['selected_seats'])
        
        if request.method == 'POST' and 'confirm_booking' in request.POST:
            passenger_form = PassengerInfoForm(request.POST, seat_count=seat_count, bus=bus)
            if passenger_form.is_valid():
                start_stop = passenger_form.cleaned_data['start_stop'].stop
                end_stop = passenger_form.cleaned_data['end_stop'].stop
                
                for i in range(seat_count):
                    seat = Seat.objects.get(pk=request.session['selected_seats'][i])
                    Booking.objects.create(
                        bus=bus,
                        customer=request.user,
                        seat=seat,
                        start_stop=start_stop,
                        end_stop=end_stop,
                        travel_date=travel_date,
                        passenger_name=passenger_form.cleaned_data[f'passenger_name_{i}'],
                        passenger_email=passenger_form.cleaned_data[f'passenger_email_{i}'],
                        status='Confirmed'
                    )
                
                del request.session['selected_seats']
                del request.session['travel_date']
                messages.success(request, 'Booking confirmed successfully!')
                return redirect('booked-buses', username=request.user.username)
            
        else:
            passenger_form = PassengerInfoForm(seat_count=seat_count, bus=bus)
        
        context = {
            'bus': bus,
            'form': passenger_form,
            'seat_count': seat_count,
            'travel_date': travel_date,
        }
        return render(request, 'bookbus/passenger_info.html', context)
        
    except Exception as e:
        messages.error(request, f'Error creating booking: {str(e)}')
        return redirect('bus-book', pk=bus.pk)

def cancel_booking(request, pk):
    booking = get_object_or_404(Booking, pk=pk, customer=request.user)
    profile = request.user.profile
    
    if request.method == 'POST':
        if booking.status != 'Confirmed':
            messages.error(request, 'Only confirmed bookings can be cancelled')
            return redirect('booked-buses', username=request.user.username)
        
        # Get current fare from the seat
        refund_amount = booking.seat.fare
        
        with transaction.atomic():
            # Update booking status
            booking.status = 'Cancelled'
            booking.cancelled_at = timezone.now()  # Add this field to track when cancelled
            booking.save()
            
            # Refund coins
            profile.coins += refund_amount
            profile.save()
            
            messages.success(request, f'Booking cancelled successfully! {refund_amount} coins refunded.')
            return redirect('booked-buses', username=request.user.username)
    
    return render(request, 'bookbus/cancel_booking.html', {'booking': booking})


def add_stop(request):
    if request.method == "POST":
        form = StopForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('bookbus-home')
    else:
        form = StopForm()
    
    return render(request, 'bookbus/add_stop.html', {'form': form})

class ExportBookingsView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    template_name = 'bookbus/export_bookings.html'
    
    def test_func(self):
        """Check if user is in BusAdmin group"""
        return self.request.user.groups.filter(name='BusAdmin').exists()
    
    def handle_no_permission(self):
        return HttpResponseForbidden("You don't have permission to access this page")

class ExportBookingsExcelView(LoginRequiredMixin, UserPassesTestMixin, View):
    def test_func(self):
        return self.request.user.groups.filter(name='BusAdmin').exists()
    
    def handle_no_permission(self):
        return HttpResponse('Unauthorized', status=401)
    
    def get(self, request, *args, **kwargs):
        # Get all bookings for buses owned by this admin (including cancelled)
        bookings = Booking.objects.filter(
            bus__travels=request.user
        ).select_related(
            'bus',
            'seat',
            'customer',
            'start_stop',
            'end_stop'
        ).order_by('bus__id', 'travel_date')

        # Create Excel workbook
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="all_bookings_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'

        wb = Workbook()
        wb.remove(wb.active)

        # Define column headers and widths
        headers = [
            ("Booking ID", 10),
            ("Bus ID", 10),
            ("Status", 12),
            ("Seat", 10),
            ("Seat Class", 15),
            ("From", 25),
            ("To", 25),
            ("Travel Date", 15),
            ("Passenger", 25),
            ("Email", 30),
            ("Fare", 10),
            ("Booked At", 20),
            ("Cancelled At", 20)
        ]

        # Group bookings by bus
        from itertools import groupby
        buses_data = {}
        for bus, bus_bookings in groupby(bookings, key=lambda x: x.bus):
            buses_data[bus] = list(bus_bookings)

        for bus, bus_bookings in buses_data.items():
            # Create worksheet for each bus
            ws = wb.create_sheet(title=f"Bus-{bus.id}")
            
            # Add bus info
            ws.append([f"Bus ID: {bus.id}"])
            ws.append([f"Route: {bus.journey_start.name} to {bus.journey_end.name}"])
            ws.append([f"Total Bookings: {len(bus_bookings)}"])
            ws.append([])
            
            # Add headers
            ws.append([h[0] for h in headers])
            
            # Style headers
            bold_font = Font(bold=True)
            for col_num, header in enumerate(headers, 1):
                col_letter = get_column_letter(col_num)
                ws.column_dimensions[col_letter].width = header[1]
                ws[f"{col_letter}{5}"].font = bold_font
            
            # Add booking data with status indication
            for booking in bus_bookings:
                row = [
                    booking.id,
                    bus.id,
                    booking.status,
                    booking.seat.name,
                    booking.seat.seat_class,
                    booking.start_stop.name,
                    booking.end_stop.name,
                    booking.travel_date.strftime("%Y-%m-%d"),
                    booking.passenger_name,
                    booking.passenger_email,
                    booking.seat.fare,
                    booking.created_at.strftime("%Y-%m-%d %H:%M"),
                    booking.cancelled_at.strftime("%Y-%m-%d %H:%M") if booking.cancelled_at else "N/A"
                ]
                ws.append(row)
            
            # Add summary
            last_row = ws.max_row
            if last_row > 5:
                ws.append([])
                ws.append(["Confirmed Bookings:", f'=COUNTIF(C6:C{last_row},"Confirmed")'])
                ws.append(["Cancelled Bookings:", f'=COUNTIF(C6:C{last_row},"Cancelled")'])
                ws.append(["Total Revenue:", f'=SUMIF(C6:C{last_row},"Confirmed",K6:K{last_row})'])
                
                # Style summary
                for cell in ['A', 'B', 'C']:
                    ws[f"{cell}{last_row+2}"].font = Font(bold=True)
                    ws[f"{cell}{last_row+3}"].font = Font(bold=True)
                    ws[f"{cell}{last_row+4}"].font = Font(bold=True)

        wb.save(response)
        return response


def about(request):
    return render(request, 'bookbus/about.html', {'title': 'About'})